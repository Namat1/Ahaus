import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from datetime import datetime

# Deutsche Monatsnamen
german_months = {
    1: "Januar", 2: "Februar", 3: "März", 4: "April",
    5: "Mai", 6: "Juni", 7: "Juli", 8: "August",
    9: "September", 10: "Oktober", 11: "November", 12: "Dezember"
}

def get_month_year(date):
    if pd.isna(date):
        return None, None
    try:
        if not isinstance(date, pd.Timestamp):
            date = pd.to_datetime(date, errors='coerce')
        if pd.isna(date):
            return None, None
        return date.month, date.year
    except:
        return None, None

def get_kw(date):
    if pd.isna(date):
        return ""
    try:
        return f"KW{pd.to_datetime(date).isocalendar().week}"
    except:
        return ""

ZULAGE_KEYWORDS = [
    "ahaus",
    "borkholzhausen",
    "glandorf",
    "optifair",
    "opti fair",
    "edv",
    "edv fleisch",
    "elfering",
    "elfering ahaus"
]

def check_zulage(comment):
    if isinstance(comment, str):
        comment_lower = comment.lower()
        return any(k in comment_lower for k in ZULAGE_KEYWORDS)
    return False


def process_file(file):
    df = pd.read_excel(file, sheet_name=0, header=None)
    df = df.iloc[4:]
    df.columns = range(df.shape[1])

    entries = []

    for _, row in df.iterrows():
        nachname_raw = row[3] if pd.notna(row[3]) else row[6]
        vorname = row[4] if pd.notna(row[4]) else row[7]
        lkw = row[11]
        datum = row[14]
        kommentar = row[15]

        if pd.notna(nachname_raw) and check_zulage(kommentar):
            monat, jahr = get_month_year(datum)
            if monat and jahr:
                nachname_check = str(nachname_raw).strip().lower()
                zulage = 0 if "zippel" in nachname_check else 20
                name = f"{nachname_raw}, {vorname}"

                ahaus_info = str(row[1]) if pd.notna(row[1]) else ""


                eintrag = {
                    "Name": name,
                    "LKW": lkw,
                    "Datum": pd.to_datetime(datum, errors='coerce'),
                    "KW": get_kw(datum),
                    "Zulage": zulage,
                    "Monat": monat,
                    "Jahr": jahr,
                    "Info": ahaus_info
                }
                entries.append(eintrag)
    return entries

def write_excel(monatsdaten):
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    # Moderne Farbpalette
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # Hellblau
    name_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Mittelblau
    data_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Weiß
    data_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")  # Hellgrau
    total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Grün für Summen
    
    # Rahmen
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    medium_border = Border(
        left=Side(style='medium', color='1F4E78'),
        right=Side(style='medium', color='1F4E78'),
        top=Side(style='medium', color='1F4E78'),
        bottom=Side(style='medium', color='1F4E78')
    )

    for (monat, jahr) in sorted(monatsdaten.keys(), key=lambda x: (x[1], x[0])):
        daten = monatsdaten[(monat, jahr)]
        ws = wb.create_sheet(f"{german_months[monat]} {jahr}")
        daten.sort(key=lambda x: (x["Name"], x["Datum"]))
        current_row = 2
        current_name = None
        alternate_row = False
        fahrer_summe = 0  # Summe für aktuellen Fahrer

        for idx, eintrag in enumerate(daten):
            name = eintrag["Name"]
            
            # Wenn neuer Fahrer beginnt, vorherigen Fahrer abschließen mit Summenzeile
            if name != current_name:
                # Summenzeile für vorherigen Fahrer (falls vorhanden)
                if current_name is not None and fahrer_summe > 0:
                    ws.cell(row=current_row, column=1, value="Gesamtzulage")
                    ws.cell(row=current_row, column=5, value=fahrer_summe)
                    
                    for col in range(1, 7):
                        cell = ws.cell(row=current_row, column=col)
                        cell.fill = total_fill
                        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.border = medium_border
                        
                        if col == 5:
                            cell.number_format = '#,##0.00 €'
                    
                    ws.row_dimensions[current_row].height = 22
                    current_row += 1
                    fahrer_summe = 0
                
                if current_name is not None:
                    current_row += 1  # Leerzeile zwischen Personen
                
                # Header-Zeile (Name, Datum, KW, LKW, Zulage, Ahaus Info)
                ws.append(["Name", "Datum", "KW", "LKW", "Zulage (€)", "Info"])
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = header_fill
                    cell.font = Font(name="Calibri", bold=True, size=10, color="1F4E78")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                current_name = name
                alternate_row = False

            # Datenzeile mit alternierender Farbe
            fill_color = data_fill_white if alternate_row else data_fill_light
            
            # Name-Zelle (erste Spalte mit Name)
            name_cell = ws.cell(row=current_row, column=1, value=name)
            name_cell.fill = name_fill
            name_cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            name_cell.border = thin_border
            
            # Restliche Zellen
            ws.cell(row=current_row, column=2, value=eintrag["Datum"].strftime("%d.%m.%Y"))
            ws.cell(row=current_row, column=3, value=eintrag["KW"])
            ws.cell(row=current_row, column=4, value=eintrag["LKW"])
            
            # Zulage mit Währungsformat
            zulage_cell = ws.cell(row=current_row, column=5, value=eintrag['Zulage'])
            zulage_cell.number_format = '#,##0.00 €'
            if eintrag['Zulage'] > 0:
                zulage_cell.font = Font(name="Calibri", size=10, color="70AD47", bold=True)
            else:
                zulage_cell.font = Font(name="Calibri", size=10, color="2C3E50")
            
            ws.cell(row=current_row, column=6, value=eintrag.get("Ahaus Info", ""))

            # Styling für Daten-Zellen (außer Name und Zulage, die schon gestyled sind)
            for col in [2, 3, 4, 6]:
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_color
                cell.font = Font(name="Calibri", size=10, color="2C3E50")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
            
            # Zulage-Zelle bekommt auch Fill und Border
            zulage_cell.fill = fill_color
            zulage_cell.alignment = Alignment(horizontal="right", vertical="center")
            zulage_cell.border = thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            alternate_row = not alternate_row
            
            # Zur Summe hinzufügen
            fahrer_summe += eintrag['Zulage']

        # Letzte Summenzeile am Ende des Sheets
        if current_name is not None and fahrer_summe > 0:
            ws.cell(row=current_row, column=1, value="Gesamtzulage")
            ws.cell(row=current_row, column=5, value=fahrer_summe)
            
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = total_fill
                cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = medium_border
                
                if col == 5:
                    cell.number_format = '#,##0.00 €'
            
            ws.row_dimensions[current_row].height = 22

        # Spaltenbreiten mit Mindestbreiten
        column_min_widths = {
            1: 25,  # Name
            2: 18,  # Datum
            3: 12,  # KW
            4: 15,  # LKW
            5: 18,  # Zulage
            6: 30   # Ahaus Info
        }
        
        max_cols = 6
        for col in range(1, max_cols + 1):
            max_length = max(
                len(str(ws.cell(row=r, column=col).value)) if ws.cell(row=r, column=col).value else 0
                for r in range(1, ws.max_row + 1)
            )
            
            # Berechne Breite mit Puffer
            calculated_width = max_length + 4
            
            # Verwende Mindestbreite falls definiert
            min_width = column_min_widths.get(col, 12)
            adjusted_width = max(calculated_width, min_width)
            
            # Maximalbreite begrenzen
            adjusted_width = min(adjusted_width, 70)
            
            ws.column_dimensions[get_column_letter(col)].width = adjusted_width

        # Freeze Panes für bessere Navigation
        ws.freeze_panes = "A3"

    wb.save(output)
    return output

# Streamlit UI
st.title("Zulage Ahaus")

uploaded_files = st.file_uploader("Excel-Dateien hochladen", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    alle_eintraege = []
    for file in uploaded_files:
        eintraege = process_file(file)
        alle_eintraege.extend(eintraege)

    if alle_eintraege:
        monatsweise = {}
        for eintrag in alle_eintraege:
            key = (eintrag["Monat"], eintrag["Jahr"])
            monatsweise.setdefault(key, []).append(eintrag)

        excel_data = write_excel(monatsweise)
        st.download_button(
            label="📥 Excel herunterladen",
            data=excel_data.getvalue(),
            file_name="Ahaus_Auswertung.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("Keine passenden Daten gefunden.")
